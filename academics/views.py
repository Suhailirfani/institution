from django.views.generic import ListView, DetailView, CreateView, UpdateView, View
from django.http import HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.db.models import Count
from django.utils.translation import gettext_lazy as _
from django.shortcuts import render, redirect
from django.db import models

from accounts.models import User
from core.views import RoleRequiredMixin
from core.services import NotificationService
from core.utils import render_to_pdf
from .models import ClassRoom, StudentProfile, StaffProfile, AttendanceRecord, ExamResult, Subject, Exam
from django.contrib import messages
from django.views.generic import FormView
from .forms import (
    ClassRoomForm, StudentProfileForm, StaffCreationForm, StudentCreationForm, 
    StudentUpdateForm, StudentBulkImportForm, StaffUpdateForm, StaffBulkImportForm,
    SubjectForm, ExamForm, BulkExamResultForm
)
from django.forms import formset_factory
from django.db import transaction
import openpyxl
from django.contrib import messages
from django.views.generic import FormView

# --- Class Room Views ---
class ClassRoomListView(RoleRequiredMixin, ListView):
    model = ClassRoom
    template_name = "academics/classroom_list.html"
    context_object_name = 'classrooms'
    allowed_roles = [User.Roles.ADMIN, User.Roles.STAFF]

    def get_queryset(self):
        return ClassRoom.objects.select_related('academic_year').annotate(student_count=Count('students'))

class ClassRoomDetailView(RoleRequiredMixin, DetailView):
    model = ClassRoom
    template_name = "academics/classroom_detail.html"
    context_object_name = 'classroom'
    allowed_roles = [User.Roles.ADMIN, User.Roles.STAFF]
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['students'] = self.object.students.all().select_related('user')
        context['subjects'] = self.object.subjects.all()
        context['exams'] = self.object.exams.all().order_by('-date')
        return context

class ClassRoomCreateView(RoleRequiredMixin, CreateView):
    model = ClassRoom
    form_class = ClassRoomForm
    template_name = "core/generic_form.html"
    success_url = reverse_lazy('academics:classroom_list')
    allowed_roles = [User.Roles.ADMIN] # Only admin creates classes

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Add New Class")
        return context

# --- Student Views ---
class StudentListView(RoleRequiredMixin, ListView):
    model = StudentProfile
    template_name = "academics/student_list.html"
    context_object_name = 'students'
    paginate_by = 20
    allowed_roles = [User.Roles.ADMIN, User.Roles.STAFF]

    def get_queryset(self):
        return StudentProfile.objects.select_related('user', 'classroom').order_by('classroom__standard', 'user__first_name')

class StudentDetailView(RoleRequiredMixin, DetailView):
    model = StudentProfile
    template_name = "academics/student_detail.html"
    context_object_name = 'student'
    allowed_roles = [User.Roles.ADMIN, User.Roles.STAFF]

class StudentCreateView(RoleRequiredMixin, CreateView):
    model = StudentProfile
    form_class = StudentCreationForm
    template_name = "academics/student_form.html"
    success_url = reverse_lazy('academics:student_list')
    allowed_roles = [User.Roles.ADMIN]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Add New Student")
        return context

class StudentUpdateView(RoleRequiredMixin, UpdateView):
    model = StudentProfile
    form_class = StudentUpdateForm
    template_name = "academics/student_form.html"
    success_url = reverse_lazy('academics:student_list')
    allowed_roles = [User.Roles.ADMIN, User.Roles.STAFF]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Update Student Profile")
        return context

class StudentSelfDetailView(LoginRequiredMixin, DetailView):
    model = StudentProfile
    template_name = "academics/student_detail.html" # Reusing detail template
    context_object_name = 'student'

    def get_object(self, queryset=None):
        if hasattr(self.request.user, 'student_profile'):
            return self.request.user.student_profile
        return None # Should handle 404 or redirect nicely

class StudentBulkImportView(RoleRequiredMixin, FormView):
    template_name = "academics/student_import.html"
    form_class = StudentBulkImportForm
    success_url = reverse_lazy('academics:student_list')
    allowed_roles = [User.Roles.ADMIN]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Bulk Import Students")
        return context

    def form_valid(self, form):
        excel_file = self.request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # Assume header is row 1
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        created_count = 0
        errors = []
        
        for index, row in enumerate(rows, start=2):
            # Columns: First Name, Last Name, Email, DOB, Admin No, WhatsApp, Class Code, Father Name, Mother Name, Address, Blood Group
            try:
                # Pad row with None if it's shorter than expected
                row = list(row) + [None] * (11 - len(row))
                first_name, last_name, email, dob, admission_number, whatsapp_number, class_code, father_name, mother_name, address, blood_group = row[:11]
                
                if not all([first_name, email, admission_number, class_code]):
                    continue # Validating mandatory fields only

                if StudentProfile.objects.filter(admission_number=admission_number).exists():
                    errors.append(f"Row {index}: Student with admission number {admission_number} already exists.")
                    continue
                    
                if User.objects.filter(email=email).exists():
                    errors.append(f"Row {index}: User with email {email} already exists.")
                    continue

                # Find classroom (Assuming Subject Code logic, but here we need Class Code which isn't unique without Institution... 
                # For MVP, let's assume Class Code maps to Standard + Division or verify strictly)
                # Actually, earlier models didn't have strict 'code' for class. Let's assume input is Standard (e.g. "10 A")
                # Wait, ClassRoom has `standard` and `division`.
                # Simplification: Assume input is "Standard-Division" e.g. "10-A"
                if '-' in str(class_code):
                     std, div = str(class_code).split('-', 1)
                     classroom = ClassRoom.objects.filter(standard__iexact=std.strip(), division__iexact=div.strip()).first()
                else:
                     classroom = ClassRoom.objects.filter(standard__iexact=str(class_code).strip()).first()

                if not classroom:
                    errors.append(f"Row {index}: Class '{class_code}' not found.")
                    continue

                # Create User
                user = User.objects.create_user(
                    username=email,
                    email=email,
                    password="password123", # Default password, should force change later
                    first_name=first_name,
                    last_name=last_name,
                    role=User.Roles.STUDENT
                )
                
                # Create Profile
                StudentProfile.objects.create(
                    user=user,
                    admission_number=admission_number,
                    whatsapp_number=whatsapp_number,
                    date_of_birth=dob,
                    classroom=classroom,
                    father_name=father_name or "",
                    mother_name=mother_name or "",
                    address=address or "",
                    blood_group=blood_group or ""
                )
                created_count += 1
                
            except Exception as e:
                errors.append(f"Row {index}: Error processing - {str(e)}")

        if created_count > 0:
            messages.success(self.request, f"Successfully imported {created_count} students.")
        
        if errors:
            messages.warning(self.request, f"Encoutered {len(errors)} errors. Check below.")
            # Pass errors to template? Or just flash huge message.
            # Ideally context, but redirects clear context. Using messages for now.
            for err in errors[:5]: # Show first 5 errors
                 messages.error(self.request, err)
            if len(errors) > 5:
                messages.error(self.request, f"...and {len(errors)-5} more errors.")
                
        return super().form_valid(form)


class DownloadStudentImportTemplateView(RoleRequiredMixin, View):
    allowed_roles = [User.Roles.ADMIN, User.Roles.STAFF]

    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Student Import Template"
        
        # Headers
        headers = ["First Name", "Last Name", "Email", "Date of Birth (YYYY-MM-DD)", "Admission Number", "WhatsApp Number", "Class Code", "Father Name", "Mother Name", "Address", "Blood Group"]
        ws.append(headers)
        
        # Sample Row
        ws.append(["John", "Doe", "john.doe@example.com", "2010-05-15", "ADM001", "9876543210", "10-A", "John Doe Sr.", "Jane Doe", "123 Street, City", "O+"])
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
             ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=student_import_template.xlsx'
        
        wb.save(response)
        return response


# --- Staff Views ---
class StaffListView(RoleRequiredMixin, ListView):
    model = StaffProfile
    template_name = "academics/staff_list.html"
    context_object_name = 'staff_members'
    allowed_roles = [User.Roles.ADMIN]

class StaffCreateView(RoleRequiredMixin, CreateView):
    model = StaffProfile
    form_class = StaffCreationForm
    template_name = "academics/staff_form.html"
    success_url = reverse_lazy('academics:staff_list')
    allowed_roles = [User.Roles.ADMIN]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Add New Staff Member")
        return context

class StaffUpdateView(RoleRequiredMixin, UpdateView):
    model = StaffProfile
    form_class = StaffUpdateForm
    template_name = "academics/staff_form.html"
    success_url = reverse_lazy('academics:staff_list')
    allowed_roles = [User.Roles.ADMIN]
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Edit Staff")
        return context

class StaffBulkImportView(RoleRequiredMixin, FormView):
    template_name = "academics/staff_import.html"
    form_class = StaffBulkImportForm
    success_url = reverse_lazy('academics:staff_list')
    allowed_roles = [User.Roles.ADMIN]

    def form_valid(self, form):
        excel_file = form.cleaned_data['excel_file']
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            
            if not rows or len(rows) < 2:
                messages.error(self.request, "The Excel file is empty or missing headers.")
                return self.form_invalid(form)

            # Skip header
            created_count = 0
            errors = []
            
            from core.models import Institution
            # For this MVP, assign to first institution or let user pick, assuming 1 for now
            institution = Institution.objects.first()

            for index, row in enumerate(rows, start=2):
                # Columns: First Name, Last Name, Email, Mobile, Place, Department, Designation, Qualification
                try:
                    # Pad row
                    row = list(row) + [None] * (8 - len(row))
                    first_name, last_name, email, mobile, place, department, designation, qualification = row[:8]
                    
                    if not all([first_name, email, mobile, place]):
                        continue # Skip if compulsory fields are missing

                    if User.objects.filter(email=email).exists():
                        errors.append(f"Row {index}: User with email {email} already exists.")
                        continue

                    # Create User
                    user = User.objects.create_user(
                        email=email,
                        username=email, # Use email as username
                        password="defaultpassword123", # Set a default password
                        first_name=first_name,
                        last_name=last_name,
                        role=User.Roles.STAFF
                    )
                    
                    # Create Profile
                    StaffProfile.objects.create(
                        user=user,
                        institution=institution,
                        mobile_number=mobile,
                        place=place,
                        department=department or "",
                        designation=designation or "",
                        qualification=qualification or ""
                    )
                    created_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {index}: Error - {str(e)}")
            
            if created_count > 0:
                messages.success(self.request, f"Successfully imported {created_count} staff members.")
            
            if errors:
                for error in errors[:10]: # Limit errors shown
                    messages.warning(self.request, error)
                if len(errors) > 10:
                    messages.warning(self.request, f"And {len(errors) - 10} more errors.")
                    
        except Exception as e:
            messages.error(self.request, f"Error processing file: {str(e)}")
            return self.form_invalid(form)
            
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Bulk Import Staff")
        return context

class DownloadStaffImportTemplateView(RoleRequiredMixin, View):
    allowed_roles = [User.Roles.ADMIN]

    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Staff Import Template"
        
        # Headers
        headers = ["First Name", "Last Name", "Email", "Mobile", "Place", "Department", "Designation", "Qualification"]
        ws.append(headers)
        
        # Sample Row
        ws.append(["Jane", "Smith", "jane.smith@example.com", "9876543210", "New York", "Science", "Teacher", "M.Sc B.Ed"])
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=staff_import_template.xlsx'
        
        wb.save(response)
        return response

# --- Attendance Views ---
class AttendanceCreateView(RoleRequiredMixin, CreateView):
    model = AttendanceRecord
    fields = ['student', 'date', 'status']  # distinct from bulk entry for MVP simplicity
    template_name = "core/generic_form.html"
    success_url = reverse_lazy('academics:student_list') # Redirect somewhere useful
    allowed_roles = [User.Roles.ADMIN, User.Roles.STAFF]

    def form_valid(self, form):
        response = super().form_valid(form)
        # Check if absent and send alert
        if self.object.status == AttendanceRecord.Status.ABSENT:
            NotificationService.send_attendance_alert(
                student=self.object.student,
                date=self.object.date,
                status=self.object.status
            )
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Mark Student Attendance")
        return context

# --- Exam / Reports Views ---
class ProgressReportView(RoleRequiredMixin, DetailView):
    model = StudentProfile
    template_name = "academics/progress_report.html"
    context_object_name = 'student'
    allowed_roles = [User.Roles.ADMIN, User.Roles.STAFF, User.Roles.PARENT, User.Roles.STUDENT]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.object
        # Fetch all results for this student
        results = ExamResult.objects.filter(student=student).select_related('exam', 'subject')
        
        # Group by Exam
        exams_data = {}
        for res in results:
            if res.exam not in exams_data:
                exams_data[res.exam] = {
                    'exam': res.exam,
                    'results': [],
                    'total_marks': 0,
                    'max_total': 0,
                    'percentage': 0,
                    'rank': 0
                }
            exams_data[res.exam]['results'].append(res)
            exams_data[res.exam]['total_marks'] += res.marks_obtained
            exams_data[res.exam]['max_total'] += res.max_marks
            
        # Calculate Rank and Percentage
        for exam, data in exams_data.items():
            if data['max_total'] > 0:
                data['percentage'] = (data['total_marks'] / data['max_total']) * 100
                
            # Rank Calculation (Expensive but okay for small classes)
            # Get all students in class
            class_students = StudentProfile.objects.filter(classroom=exam.classroom)
            exam_totals = []
            for s in class_students:
                # Sum marks for this student in this exam
                total = ExamResult.objects.filter(exam=exam, student=s).aggregate(t=models.Sum('marks_obtained'))['t'] or 0
                exam_totals.append(total)
            
            # Sort descending
            exam_totals.sort(reverse=True)
            # Find rank
            try:
                data['rank'] = exam_totals.index(data['total_marks']) + 1
            except ValueError:
                data['rank'] = '-'
            
        context['exams_data'] = exams_data.values()
        return context

class ClassExamResultView(RoleRequiredMixin, DetailView):
    model = Exam
    template_name = "academics/class_exam_result.html"
    context_object_name = 'exam'
    allowed_roles = [User.Roles.ADMIN, User.Roles.STAFF]
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        exam = self.object
        students = StudentProfile.objects.filter(classroom=exam.classroom).select_related('user')
        
        student_results = []
        for student in students:
            # Aggregate marks
            results = ExamResult.objects.filter(exam=exam, student=student)
            total_marks = results.aggregate(t=models.Sum('marks_obtained'))['t'] or 0
            max_marks = results.aggregate(t=models.Sum('max_marks'))['t'] or 0
            
            percentage = (total_marks / max_marks * 100) if max_marks > 0 else 0
            
            student_results.append({
                'student': student,
                'total_marks': total_marks,
                'max_marks': max_marks,
                'percentage': percentage,
            })
            
        # Sort by total marks desc
        student_results.sort(key=lambda x: x['total_marks'], reverse=True)
        
        # Add Rank
        for i, item in enumerate(student_results):
            item['rank'] = i + 1
            
        context['student_results'] = student_results
        context['toppers'] = student_results[:5] # Top 5
        return context

class StudentCertificateView(RoleRequiredMixin, DetailView):
    model = StudentProfile
    template_name = "academics/student_certificate.html"
    context_object_name = 'student'
    allowed_roles = [User.Roles.ADMIN, User.Roles.STAFF]
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Try to get the institution from the student's classroom, or fallback to the first active one
        if self.object.classroom and self.object.classroom.institution:
             context['institution'] = self.object.classroom.institution
        else:
             from core.models import Institution
             context['institution'] = Institution.objects.filter(is_active=True).first()
        return context

# --- ID Card Views ---
class StudentIDCardView(RoleRequiredMixin, DetailView):
    model = StudentProfile
    template_name = "academics/pdf/student_id.html"
    context_object_name = 'student'
    allowed_roles = [User.Roles.ADMIN, User.Roles.STAFF, User.Roles.STUDENT]

    def render_to_response(self, context, **response_kwargs):
        # We need a list for the template even for single item to reuse template logic
        context['students'] = [self.object]
        pdf = render_to_pdf(self.template_name, context)
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = f"Identity_Card_{self.object.admission_number}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        return HttpResponse("Error generating PDF", status=500)

class StudentBulkIDCardView(RoleRequiredMixin, View):
    allowed_roles = [User.Roles.ADMIN, User.Roles.STAFF]

    def get(self, request, *args, **kwargs):
        # Optional: Filter by Class
        class_id = request.GET.get('class_id')
        if class_id:
            students = StudentProfile.objects.filter(classroom_id=class_id).select_related('user', 'classroom', 'classroom__institution')
        else:
            students = StudentProfile.objects.all().select_related('user', 'classroom', 'classroom__institution')

        if not students.exists():
            messages.warning(request, "No students found to generate ID cards.")
            return HttpResponse("No students found", status=404)

        context = {'students': students}
        pdf = render_to_pdf("academics/pdf/student_id.html", context)
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "Student_ID_Cards_Bulk.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        return HttpResponse("Error generating PDF", status=500)

class StaffIDCardView(RoleRequiredMixin, DetailView):
    model = StaffProfile
    template_name = "academics/pdf/staff_id.html"
    context_object_name = 'staff'
    allowed_roles = [User.Roles.ADMIN, User.Roles.STAFF] # Staff can download their own

    def render_to_response(self, context, **response_kwargs):
        context['staff_members'] = [self.object]
        pdf = render_to_pdf(self.template_name, context)
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = f"Staff_Identity_Card_{self.object.user.username}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        return HttpResponse("Error generating PDF", status=500)

class StaffBulkIDCardView(RoleRequiredMixin, View):
    allowed_roles = [User.Roles.ADMIN]

    def get(self, request, *args, **kwargs):
        staff_members = StaffProfile.objects.all().select_related('user', 'institution')
        
        if not staff_members.exists():
             messages.warning(request, "No staff members found.")
             return HttpResponse("No staff found", status=404)

        context = {'staff_members': staff_members}
        pdf = render_to_pdf("academics/pdf/staff_id.html", context)
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "Staff_ID_Cards_Bulk.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        return HttpResponse("Error generating PDF", status=500)

# --- Result Management Views ---

class SubjectListView(RoleRequiredMixin, ListView):
    model = Subject
    template_name = "components/data_table.html"
    context_object_name = 'subjects'
    allowed_roles = [User.Roles.ADMIN, User.Roles.STAFF]
    
    def get_queryset(self):
        # Filter by classroom if provided
        classroom_id = self.request.GET.get('classroom')
        if classroom_id:
             return Subject.objects.filter(classroom_id=classroom_id).select_related('classroom')
        return Subject.objects.all().select_related('classroom')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['table_title'] = _("Subjects")
        context['columns'] = [
             {'header': 'Code', 'field': 'code'},
             {'header': 'Name', 'field': 'name'},
             {'header': 'Classroom', 'field': 'classroom'},
             {'header': 'Max Marks', 'field': 'max_marks'},
             {'header': 'Pass Marks', 'field': 'pass_marks'},
             {'header': 'Action', 'field': 'action'},
        ]
        self.template_name = "academics/subject_list.html"
        return context

class SubjectCreateView(RoleRequiredMixin, CreateView):
    model = Subject
    form_class = SubjectForm
    template_name = "core/generic_form.html"
    success_url = reverse_lazy('academics:subject_list') # Redirects to list, better if to class detail
    allowed_roles = [User.Roles.ADMIN]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Add Subject")
        return context
        
    def get_success_url(self):
        classroom_id = self.object.classroom.id
        return reverse_lazy('academics:classroom_detail', kwargs={'pk': classroom_id})

class ExamListView(RoleRequiredMixin, ListView):
    model = Exam
    template_name = "components/data_table.html"
    context_object_name = 'exams'
    allowed_roles = [User.Roles.ADMIN, User.Roles.STAFF]
    
    def get_queryset(self):
        classroom_id = self.request.GET.get('classroom')
        if classroom_id:
             return Exam.objects.filter(classroom_id=classroom_id).select_related('classroom', 'academic_year')
        return Exam.objects.all().select_related('classroom', 'academic_year')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['table_title'] = _("Exams")
        context['columns'] = [
             {'header': 'Name', 'field': 'name'},
             {'header': 'Classroom', 'field': 'classroom'},
             {'header': 'Date', 'field': 'date'},
             {'header': 'Action', 'field': 'action'},
        ]
        self.template_name = "academics/exam_list.html"
        return context

class ExamCreateView(RoleRequiredMixin, CreateView):
    model = Exam
    form_class = ExamForm
    template_name = "core/generic_form.html"
    allowed_roles = [User.Roles.ADMIN]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Schedule Exam")
        return context

    def get_success_url(self):
        return reverse_lazy('academics:classroom_detail', kwargs={'pk': self.object.classroom.pk})

# Data Entry View
class ExamResultEntryView(RoleRequiredMixin, View):
    allowed_roles = [User.Roles.ADMIN, User.Roles.STAFF]
    template_name = "academics/result_entry.html"

    def get(self, request, exam_id, subject_id):
        exam = Exam.objects.get(pk=exam_id)
        subject = Subject.objects.get(pk=subject_id)
        
        # Get all students in the exam's classroom
        students = StudentProfile.objects.filter(classroom=exam.classroom).order_by('admission_number')
        
        # Get existing results to pre-fill
        existing_results = ExamResult.objects.filter(exam=exam, subject=subject).in_bulk(field_name='student_id')
        
        initial_data = []
        for student in students:
            result = existing_results.get(student.id)
            initial_data.append({
                'student_id': student.id,
                'student_name': student.user.get_full_name() or student.admission_number,
                'marks_obtained': result.marks_obtained if result else None,
                'is_absent': False, # Can elaborate this later if we track absent state explicitly in DB
            })
            
        BulkResultFormSet = formset_factory(BulkExamResultForm, extra=0)
        # We need to manually populate the formset initial data because it's a bit tricky with formsets
        # But let's try passing initial to formset_factory result instantiation
        formset = BulkResultFormSet(initial=initial_data)
        
        context = {
            'exam': exam,
            'subject': subject,
            'formset': formset,
        }
        return render(request, self.template_name, context)

    def post(self, request, exam_id, subject_id):
        exam = Exam.objects.get(pk=exam_id)
        subject = Subject.objects.get(pk=subject_id)
        
        BulkResultFormSet = formset_factory(BulkExamResultForm)
        formset = BulkResultFormSet(request.POST) 
        
        if formset.is_valid():
            with transaction.atomic():
                for form in formset:
                    data = form.cleaned_data
                    if not data: continue # Empty form
                    
                    student_id = data.get('student_id')
                    marks = data.get('marks_obtained')
                    is_absent = data.get('is_absent')
                    
                    if student_id:
                        student = StudentProfile.objects.get(pk=student_id)
                        
                        if marks is not None:
                            ExamResult.objects.update_or_create(
                                exam=exam,
                                student=student,
                                subject=subject,
                                defaults={
                                    'marks_obtained': marks,
                                    'max_marks': subject.max_marks, # Default to subject's max marks
                                    'grade': '', # Calculate grade later if needed
                                }
                            )
            messages.success(request, _("Marks saved successfully."))
            return redirect('academics:classroom_detail', pk=exam.classroom.pk)
            
        context = {
            'exam': exam,
            'subject': subject,
            'formset': formset,
        }
        return render(request, self.template_name, context)


class StudentCertificateView(RoleRequiredMixin, DetailView):
    model = StudentProfile
    template_name = "academics/student_certificate.html"
    context_object_name = 'student'
    allowed_roles = [User.Roles.ADMIN, User.Roles.STAFF, User.Roles.STUDENT, User.Roles.PARENT] # Accessible to students too

    def render_to_response(self, context, **response_kwargs):
        pdf = render_to_pdf(self.template_name, context)
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = f"Certificate_{self.object.admission_number}.pdf"
            response['Content-Disposition'] = f'inline; filename="{filename}"'
            return response
        return HttpResponse("Error generating PDF", status=500)


class StudentMarksheetPDFView(RoleRequiredMixin, DetailView):
    model = StudentProfile
    template_name = "academics/pdf/marksheet.html"
    context_object_name = 'student'
    allowed_roles = [User.Roles.ADMIN, User.Roles.STAFF, User.Roles.STUDENT, User.Roles.PARENT]

    def render_to_response(self, context, **response_kwargs):
        # We need to populate context with exam data same as ProgressReportView
        student = self.object
        results = ExamResult.objects.filter(student=student).select_related('exam', 'subject')
        
        exams_data = {}
        for res in results:
            if res.exam not in exams_data:
                exams_data[res.exam] = {
                    'exam': res.exam,
                    'results': [],
                    'total_marks': 0,
                    'max_total': 0,
                    'percentage': 0,
                    'rank': 0
                }
            exams_data[res.exam]['results'].append(res)
            exams_data[res.exam]['total_marks'] += res.marks_obtained
            exams_data[res.exam]['max_total'] += res.max_marks
            
        for exam, data in exams_data.items():
            if data['max_total'] > 0:
                data['percentage'] = (data['total_marks'] / data['max_total']) * 100
            
            # Rank Calculation (simplified for PDF view, separate query for each student)
            class_students = StudentProfile.objects.filter(classroom=exam.classroom)
            exam_totals = []
            for s in class_students:
                 # Optimize: We could annotatate but this is distinct queries in loop
                 # For MVP stick to loop or duplicate logic
                total = ExamResult.objects.filter(exam=exam, student=s).aggregate(t=models.Sum('marks_obtained'))['t'] or 0
                exam_totals.append(total)
            exam_totals.sort(reverse=True)
            try:
                data['rank'] = exam_totals.index(data['total_marks']) + 1
            except ValueError:
                data['rank'] = '-'

        context['exams_data'] = exams_data.values()
        
        pdf = render_to_pdf(self.template_name, context)
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = f"Marksheet_{student.admission_number}.pdf"
            # Use attachment to trigger download instead of inline preview
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        return HttpResponse("Error generating PDF", status=500)